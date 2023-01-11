from openpyxl import Workbook
from io import BytesIO
from employees.models import Employee
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, Protection



def get_excel_employees(employees):
        excelfile = BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet = workbook.create_sheet(title='Employees and salary', index=1)

        columns = ['Name', 'Salary']
        row_num = 1
     
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)
        # Iterate through all coins
        for _, employee in enumerate(employees, 1):
            row_num += 1

         # Define the data for each cell in the row
            row = [
            employee.name,
            employee.salary,
            ]

        # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.protection = Protection(locked=True)
        # for row in employees:
        #     worksheet.append(row)

        workbook.save(excelfile)

        # Create the HttpResponse object with the appropriate Excel headers
        response = HttpResponse(excelfile.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = 'attachment; filename=myfile.xlsx'
        return response

